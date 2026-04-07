"""Import engine: parse files, validate, detect duplicates, commit."""
from decimal import Decimal
from difflib import get_close_matches
from uuid import UUID
import io
import csv

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models import ImportSession, Voucher, VoucherEntry, Ledger, Company
from utils.tally_parser import parse_tally_xml, parse_tally_masters
from services.marg_parser import parse_marg_csv as _parse_marg_csv
from services.excel_parser import (
    suggest_column_mapping,
    parse_excel_with_mapping,
    parse_csv_with_mapping,
    get_excel_sheet_names,
)
from services.bank_statement_parser import parse_bank_statement


def parse_tally_xml_content(content: str | bytes) -> list[dict]:
    """Parse Tally XML to normalized vouchers (dict list). Includes bill_ref and gst_type from entries."""
    vouchers = parse_tally_xml(content)
    out = []
    for v in vouchers:
        out.append({
            "voucher_type": v.voucher_type,
            "date": v.date,
            "narration": v.narration,
            "reference": v.reference,
            "party_ledger_name": v.party_ledger,
            "amount": float(v.amount) if v.amount else None,
            "tally_guid": v.tally_guid,
            "entries": [
                {
                    "ledger_name": e.ledger_name,
                    "dr_amount": e.dr_amount,
                    "cr_amount": e.cr_amount,
                    "narration": e.narration,
                    "bill_ref": e.bill_ref or None,
                    "gst_type": e.gst_type or None,
                }
                for e in v.entries
            ],
            "inventory_lines": v.inventory_lines,
        })
    return out


def parse_tally_masters_content(content: str | bytes) -> list[dict]:
    """Parse Tally XML to ledger masters (dict list)."""
    masters = parse_tally_masters(content)
    return [
        {
            "name": m.name,
            "parent": m.parent,
            "opening_balance": m.opening_balance,
            "gstn": m.gstn,
            "country": m.country,
        }
        for m in masters
    ]


def parse_marg_csv(content: str | bytes) -> list[dict]:
    """Parse Marg CSV using core/services/marg_parser."""
    return _parse_marg_csv(content)


def parse_excel(content: bytes, column_mapping: dict[str, str] | None = None, sheet_index: int = 0) -> list[dict]:
    """Parse Excel; if column_mapping omitted, auto-suggest from first sheet headers."""
    if column_mapping is None:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb[wb.sheetnames[sheet_index]] if wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            return []
        if not rows:
            return []
        headers = [str(c).strip() if c else "" for c in rows[0]]
        column_mapping = suggest_column_mapping(headers)
    return parse_excel_with_mapping(content, column_mapping, sheet_index)


def get_excel_sheets(content: bytes) -> list[str]:
    """Return sheet names for column mapping UI."""
    return get_excel_sheet_names(content)


def parse_bank_statement_content(content: str | bytes) -> tuple[str, list[dict]]:
    """Parse bank statement CSV; returns (bank_key, list of {date, description, debit, credit, balance})."""
    return parse_bank_statement(content)


def fuzzy_match_ledger(name: str, ledger_names: list[str], cutoff: float = 0.6) -> str | None:
    """Return best matching ledger name from list, or None."""
    if not name or not ledger_names:
        return None
    name_upper = name.strip().upper()
    if name_upper in [n.strip().upper() for n in ledger_names]:
        return name
    matches = get_close_matches(name_upper, [n.strip().upper() for n in ledger_names], n=1, cutoff=cutoff)
    if matches:
        idx = [n.strip().upper() for n in ledger_names].index(matches[0])
        return ledger_names[idx]
    return None


def detect_duplicates(
    vouchers: list[dict],
    keys: list[str] | None = None,
    existing_guids: set[str] | None = None,
) -> list[dict]:
    """Detect duplicates by tally_guid (using existing_guids + in-file), or by keys (date, reference, amount)."""
    duplicates = []
    if existing_guids is not None or any(v.get("tally_guid") for v in vouchers):
        seen_guid = set(existing_guids) if existing_guids else set()
        for i, v in enumerate(vouchers):
            g = v.get("tally_guid")
            if g:
                if g in seen_guid:
                    duplicates.append({"index": i, "duplicate_of": "in_file", "voucher": v, "guid": g})
                elif existing_guids and g in existing_guids:
                    duplicates.append({"index": i, "duplicate_of": "existing", "voucher": v, "guid": g})
                seen_guid.add(g)
        if duplicates:
            return duplicates
    keys = keys or ["date", "reference", "amount"]
    seen = {}
    for i, v in enumerate(vouchers):
        k = tuple(v.get(k) for k in keys if k in v)
        if k in seen:
            duplicates.append({"index": i, "duplicate_of": seen[k], "voucher": v})
        else:
            seen[k] = i
    return duplicates


async def validate_import(
    db: AsyncSession,
    company_id: UUID,
    vouchers: list[dict],
    fuzzy_ledger: bool = True,
) -> tuple[list[dict], list[dict]]:
    """Validate vouchers: Dr=Cr, ledger names exist (exact or fuzzy match). Returns (valid[], errors[])."""
    valid = []
    errors = []
    result = await db.execute(select(Ledger).where(Ledger.company_id == company_id))
    ledgers = result.scalars().all()
    ledger_by_name = {l.name.strip().upper(): l.id for l in ledgers}
    ledger_names = [l.name for l in ledgers]

    for i, v in enumerate(vouchers):
        errs = []
        entries = v.get("entries", [])
        total_dr = sum(Decimal(str(e.get("dr_amount", 0))) for e in entries)
        total_cr = sum(Decimal(str(e.get("cr_amount", 0))) for e in entries)
        if total_dr != total_cr:
            errs.append("Dr != Cr")
        resolved = []
        for e in entries:
            name = (e.get("ledger_name") or "").strip()
            name_upper = name.upper()
            if not name:
                continue
            if name_upper in ledger_by_name:
                resolved.append((e, ledger_by_name[name_upper]))
            elif fuzzy_ledger:
                match = fuzzy_match_ledger(name, ledger_names)
                if match and match.upper() in ledger_by_name:
                    resolved.append((e, ledger_by_name[match.upper()]))
                    e["resolved_ledger_id"] = ledger_by_name[match.upper()]
                else:
                    errs.append(f"Ledger not found: {e.get('ledger_name')}")
            else:
                errs.append(f"Ledger not found: {e.get('ledger_name')}")
        if errs:
            errors.append({"index": i, "errors": errs, "voucher": v})
        else:
            valid.append(v)
    return valid, errors


async def commit_import(
    db: AsyncSession,
    session_id: UUID,
) -> dict:
    """Mark import session committed (vouchers already inserted by router). Update session status."""
    result = await db.execute(select(ImportSession).where(ImportSession.id == session_id))
    sess = result.scalar_one_or_none()
    if not sess:
        return {"ok": False, "message": "Session not found"}
    sess.status = "completed"
    await db.flush()
    return {"ok": True, "imported_records": sess.imported_records}
