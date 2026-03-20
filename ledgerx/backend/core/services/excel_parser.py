"""
Excel/CSV smart parser: auto-detect column headers (fuzzy match), map to standard voucher structure.
Supports multi-sheet Excel; user can confirm column mapping via UI.
"""
import csv
import io
import re
from decimal import Decimal
from typing import Any


# Standard target fields for voucher mapping
STANDARD_FIELDS = [
    "date", "voucher_type", "voucher_number", "party_ledger", "narration", "reference",
    "ledger", "debit", "credit", "amount", "dr_cr",
    "item", "qty", "rate", "batch", "expiry",
]


def _normalize_header(h: str) -> str:
    return re.sub(r"[_\s]+", "", (h or "").lower())


def _fuzzy_match(a: str, b: str) -> float:
    """Simple similarity 0-1. Exact match = 1."""
    a, b = _normalize_header(a), _normalize_header(b)
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.9
    # Jaccard-ish on words
    sa, sb = set(a), set(b)
    return len(sa & sb) / len(sa | sb) if (sa | sb) else 0.0


def suggest_column_mapping(headers: list[str]) -> dict[str, str]:
    """
    Given list of column headers from file, suggest mapping to standard fields.
    Returns dict: standard_field -> column_name from file.
    """
    mapping: dict[str, str] = {}
    aliases = {
        "date": ["date", "vchdate", "dt", "transactiondate", "voucherdate"],
        "voucher_type": ["vchtype", "vouchertype", "type", "vtype"],
        "voucher_number": ["vchno", "number", "vouchernumber", "ref", "no"],
        "party_ledger": ["party", "partyledger", "ledgername", "account", "party name"],
        "narration": ["narration", "narr", "remarks", "particulars", "description"],
        "reference": ["reference", "ref", "billref"],
        "ledger": ["ledger", "ledgername", "account"],
        "debit": ["debit", "dr", "debitamount"],
        "credit": ["credit", "cr", "creditamount"],
        "amount": ["amount", "amt", "value"],
        "dr_cr": ["type", "drcr", "dr/cr", "debitcredit"],
        "item": ["item", "stockitem", "product", "itemname"],
        "qty": ["qty", "quantity", "qty"],
        "rate": ["rate", "price", "unitprice"],
        "batch": ["batch", "batchno"],
        "expiry": ["expiry", "expirydate", "exp date"],
    }
    for std in STANDARD_FIELDS:
        best_score = 0.0
        best_col = None
        for col in headers:
            cnorm = _normalize_header(col)
            for alias in aliases.get(std, [std]):
                score = _fuzzy_match(col, alias)
                if score > best_score:
                    best_score = score
                    best_col = col
        if best_col and best_score >= 0.5:
            mapping[std] = best_col
    return mapping


def parse_excel_with_mapping(
    content: bytes,
    column_mapping: dict[str, str],
    sheet_index: int = 0,
) -> list[dict[str, Any]]:
    """
    Read Excel (or CSV) and parse rows using column_mapping (standard_field -> column name in file).
    column_mapping: e.g. {"date": "Date", "ledger": "Ledger", "debit": "Debit", "credit": "Credit"}
    Returns list of normalized voucher dicts.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        sheets = wb.sheetnames
        if sheet_index >= len(sheets):
            sheet_index = 0
        ws = wb[sheets[sheet_index]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return []

    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    rev_map = {v: k for k, v in column_mapping.items() if v in headers}
    out = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        std = {}
        for col_name, std_name in rev_map.items():
            val = d.get(col_name)
            if val is not None and str(val).strip():
                std[std_name] = val
        if std:
            # Normalize to voucher entry format
            date_val = str(std.get("date", "")).strip()
            vtype = str(std.get("voucher_type", "Journal")).strip()
            vno = str(std.get("voucher_number", "")).strip()
            ledger = str(std.get("ledger", "") or std.get("party_ledger", "")).strip()
            dr = std.get("debit")
            cr = std.get("credit")
            amt = std.get("amount")
            if dr is not None or cr is not None:
                try:
                    dr_val = Decimal(str(dr).replace(",", "")) if dr else Decimal("0")
                except Exception:
                    dr_val = Decimal("0")
                try:
                    cr_val = Decimal(str(cr).replace(",", "")) if cr else Decimal("0")
                except Exception:
                    cr_val = Decimal("0")
            elif amt is not None:
                try:
                    a = Decimal(str(amt).replace(",", ""))
                except Exception:
                    a = Decimal("0")
                drcr = str(std.get("dr_cr", "Dr")).upper()
                if "CR" in drcr or drcr == "C":
                    dr_val, cr_val = Decimal("0"), a
                else:
                    dr_val, cr_val = a, Decimal("0")
            else:
                dr_val, cr_val = Decimal("0"), Decimal("0")

            out.append({
                "date": date_val,
                "voucher_type": vtype,
                "voucher_number": vno,
                "narration": str(std.get("narration", "")).strip(),
                "reference": str(std.get("reference", "")).strip(),
                "party_ledger_name": str(std.get("party_ledger", "")).strip(),
                "entries": [{"ledger_name": ledger, "dr_amount": dr_val, "cr_amount": cr_val, "narration": str(std.get("narration", ""))}],
                "inventory_lines": [{"stock_item": str(std.get("item", "")), "quantity": Decimal(str(std.get("qty", 0) or 0)), "rate": Decimal(str(std.get("rate", 0) or 0))}] if std.get("item") else [],
            })
    return _group_into_vouchers(out)


def _group_into_vouchers(rows: list[dict]) -> list[dict]:
    """Group rows by date+voucher_number into single vouchers with multiple entries."""
    by_key: dict[tuple[str, str], dict] = {}
    for r in rows:
        key = (r.get("date", ""), r.get("voucher_number", ""))
        if key not in by_key:
            by_key[key] = {
                "date": r.get("date", ""),
                "voucher_type": r.get("voucher_type", "Journal"),
                "voucher_number": r.get("voucher_number", ""),
                "narration": r.get("narration", ""),
                "reference": r.get("reference", ""),
                "party_ledger_name": r.get("party_ledger_name", ""),
                "entries": [],
                "inventory_lines": [],
            }
        by_key[key]["entries"].extend(r.get("entries", []))
        by_key[key]["inventory_lines"].extend(r.get("inventory_lines", []))
    return list(by_key.values())


def parse_csv_with_mapping(content: str | bytes, column_mapping: dict[str, str]) -> list[dict[str, Any]]:
    """Parse CSV with given column mapping. Returns same format as parse_excel_with_mapping."""
    if isinstance(content, bytes):
        content = content.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    rev_map = {v: k for k, v in column_mapping.items() if v in headers}
    out = []
    for row in reader:
        std = {}
        for col_name, std_name in rev_map.items():
            val = row.get(col_name)
            if val is not None and str(val).strip():
                std[std_name] = val
        if not std:
            continue
        ledger = str(std.get("ledger", "") or std.get("party_ledger", "")).strip()
        dr = std.get("debit")
        cr = std.get("credit")
        amt = std.get("amount")
        if dr is not None or cr is not None:
            try:
                dr_val = Decimal(str(dr).replace(",", "")) if dr else Decimal("0")
            except Exception:
                dr_val = Decimal("0")
            try:
                cr_val = Decimal(str(cr).replace(",", "")) if cr else Decimal("0")
            except Exception:
                cr_val = Decimal("0")
        elif amt is not None:
            try:
                a = Decimal(str(amt).replace(",", ""))
            except Exception:
                a = Decimal("0")
            drcr = str(std.get("dr_cr", "Dr")).upper()
            dr_val, cr_val = (Decimal("0"), a) if ("CR" in drcr or drcr == "C") else (a, Decimal("0"))
        else:
            dr_val, cr_val = Decimal("0"), Decimal("0")

        out.append({
            "date": str(std.get("date", "")),
            "voucher_type": str(std.get("voucher_type", "Journal")),
            "voucher_number": str(std.get("voucher_number", "")),
            "narration": str(std.get("narration", "")),
            "reference": str(std.get("reference", "")),
            "party_ledger_name": str(std.get("party_ledger", "")),
            "entries": [{"ledger_name": ledger, "dr_amount": dr_val, "cr_amount": cr_val, "narration": str(std.get("narration", ""))}],
            "inventory_lines": [],
        })
    return _group_into_vouchers(out)


def get_excel_sheet_names(content: bytes) -> list[str]:
    """Return list of sheet names for Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []
